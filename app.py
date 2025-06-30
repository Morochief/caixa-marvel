from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

print("🚀 Ejecutando app.py desde:", __file__)  # Verificar qué archivo estás ejecutando

app = Flask(__name__)
CORS(app)

# Configuración SQLite
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///caixa.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# MODELOS
class Registro(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.String(20))
    frota = db.Column(db.String(100))
    chofer = db.Column(db.String(100))
    concepto = db.Column(db.String(200))
    lugar = db.Column(db.String(200))
    monto = db.Column(db.Float)
    saldo = db.Column(db.Float)

class Saldo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    valor = db.Column(db.Float)

# ARCHIVOS ORIGINALES
DATA_FILE = 'data.json'
SALDO_FILE = 'saldo_inicial.json'

# MIGRAR DATOS SI EXISTEN
with app.app_context():
    db.create_all()
    print("✔ Base de datos creada o ya existente")  # Confirmación visual

    # Migrar saldo inicial
    if not Saldo.query.first() and os.path.exists(SALDO_FILE):
        with open(SALDO_FILE, 'r', encoding='utf-8') as f:
            valor = json.load(f).get("saldo_inicial", 0)
            db.session.add(Saldo(valor=valor))
            db.session.commit()

    # Migrar registros
    if Registro.query.count() == 0 and os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            registros = json.load(f)
            saldo_actual = Saldo.query.first().valor if Saldo.query.first() else 0
            for r in registros:
                monto = float(r.get("monto", 0))
                saldo_actual += monto
                db.session.add(Registro(
                    fecha=r.get("fecha"),
                    frota=r.get("frota"),
                    chofer=r.get("chofer"),
                    concepto=r.get("concepto"),
                    lugar=r.get("lugar"),
                    monto=monto,
                    saldo=saldo_actual
                ))
            db.session.commit()

# FUNCIONES DE FORMATO
def formato_miles_punto(n):
    try:
        return "{:,.0f}".format(float(n)).replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return n

def formato_fecha_ddmmaaaa(fecha_str):
    try:
        dt = datetime.strptime(fecha_str, "%Y-%m-%d")
        return dt.strftime("%d/%m/%Y")
    except:
        return fecha_str

# RUTAS
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/saldo', methods=['GET'])
def obtener_saldo():
    s = Saldo.query.first()
    return jsonify({'saldo_inicial': s.valor if s else 0})

@app.route('/saldo', methods=['PUT'])
def actualizar_saldo():
    nuevo_valor = request.json.get('saldo_inicial', 0)
    s = Saldo.query.first()
    if not s:
        s = Saldo(valor=nuevo_valor)
        db.session.add(s)
    else:
        s.valor = nuevo_valor
    db.session.commit()

    # Recalcular todos los saldos
    registros = Registro.query.order_by(Registro.id).all()
    saldo = nuevo_valor
    for r in registros:
        saldo += r.monto
        r.saldo = saldo
    db.session.commit()

    return jsonify({'message': 'Saldo inicial actualizado'})

@app.route('/registros', methods=['GET'])
def listar():
    return jsonify([{
        'id': r.id,
        'fecha': r.fecha,
        'frota': r.frota,
        'chofer': r.chofer,
        'concepto': r.concepto,
        'lugar': r.lugar,
        'monto': r.monto,
        'saldo': r.saldo
    } for r in Registro.query.order_by(Registro.id).all()])

@app.route('/registros', methods=['POST'])
def crear():
    data = request.json
    s = Saldo.query.first()
    saldo = s.valor if s else 0
    registros = Registro.query.order_by(Registro.id).all()
    for r in registros:
        saldo += r.monto

    monto = float(data.get('monto', 0))
    nuevo_saldo = saldo + monto

    nuevo = Registro(
        fecha=data.get('fecha'),
        frota=data.get('frota'),
        chofer=data.get('chofer'),
        concepto=data.get('concepto'),
        lugar=data.get('lugar'),
        monto=monto,
        saldo=nuevo_saldo
    )
    db.session.add(nuevo)
    db.session.commit()
    return jsonify({'message': 'Registro agregado'})

@app.route('/registros/<int:id>', methods=['PUT'])
def actualizar(id):
    data = request.json
    r = Registro.query.get_or_404(id)
    r.fecha = data.get('fecha')
    r.frota = data.get('frota')
    r.chofer = data.get('chofer')
    r.concepto = data.get('concepto')
    r.lugar = data.get('lugar')
    r.monto = float(data.get('monto', 0))
    db.session.commit()

    # Recalcular saldos
    registros = Registro.query.order_by(Registro.id).all()
    saldo = Saldo.query.first().valor
    for reg in registros:
        saldo += reg.monto
        reg.saldo = saldo
    db.session.commit()
    return jsonify({'message': 'Registro actualizado'})

@app.route('/registros/<int:id>', methods=['DELETE'])
def eliminar(id):
    r = Registro.query.get_or_404(id)
    db.session.delete(r)
    db.session.commit()

    # Recalcular saldos
    registros = Registro.query.order_by(Registro.id).all()
    saldo = Saldo.query.first().valor
    for reg in registros:
        saldo += reg.monto
        reg.saldo = saldo
    db.session.commit()
    return jsonify({'message': 'Registro eliminado'})

@app.route('/exportar', methods=['GET'])
def exportar():
    registros = Registro.query.order_by(Registro.id).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Registros CAIXA"

    headers = ['Fecha', 'Frota', 'Chofer', 'Concepto', 'Lugar', 'Monto', 'Saldo']
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1E88E5")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center")

    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

    emoji_map = {
        'paraguay': '🇵🇾',
        'brasil': '🇧🇷',
        'argentina': '🇦🇷',
        'uruguay': '🇺🇾',
        'chile': '🇨🇱'
    }

    for reg in registros:
        concepto = reg.concepto.lower()
        emoji = ''
        for pais, icono in emoji_map.items():
            if pais in concepto:
                emoji = icono
                break

        concepto_final = f"{emoji} {reg.concepto}" if emoji else reg.concepto
        fecha_formateada = formato_fecha_ddmmaaaa(reg.fecha)
        monto_formateado = formato_miles_punto(reg.monto)
        saldo_formateado = formato_miles_punto(reg.saldo)

        ws.append([
            fecha_formateada,
            reg.frota,
            reg.chofer,
            concepto_final,
            reg.lugar,
            monto_formateado,
            saldo_formateado
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        concepto_valor = str(row[3].value).upper()
        for cell in row:
            cell.font = Font(name="Calibri", size=11, color="000000")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill("solid", fgColor="E3F2FD")
            if "TRANSFERENCIA DE GUARANI-CHARLES" in concepto_valor:
                cell.fill = PatternFill("solid", fgColor="C8E6C9")

    saldo_final = registros[-1].saldo if registros else 0
    fecha_final = datetime.now().strftime("%d/%m/%Y")

    fill_amarillo = PatternFill("solid", fgColor="FFF9C4")
    ws.append([])

    fila_saldo = ws.max_row + 1
    ws.merge_cells(start_row=fila_saldo, start_column=1, end_row=fila_saldo, end_column=6)
    celda_saldo_fecha = ws.cell(row=fila_saldo, column=1)
    celda_saldo_fecha.value = f"SALDO AL CIERRE - {fecha_final}"
    celda_saldo_fecha.fill = fill_amarillo
    celda_saldo_fecha.font = Font(bold=True, size=14, color="004D40")
    celda_saldo_fecha.alignment = Alignment(horizontal="center", vertical="center")

    celda_saldo_monto = ws.cell(row=fila_saldo, column=7)
    celda_saldo_monto.value = formato_miles_punto(saldo_final)
    celda_saldo_monto.fill = fill_amarillo
    celda_saldo_monto.font = Font(bold=True, size=14, color="004D40")
    celda_saldo_monto.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    ruta_archivo = os.path.join(os.path.expanduser("~/OneDrive/Desktop"), "registros_estilizados.xlsx")
    wb.save(ruta_archivo)

    return send_file(ruta_archivo, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
